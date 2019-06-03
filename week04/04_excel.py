"""
엑셀 파일 다루기
workbook, worksheet
# pip install openpyxl
"""
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = "test"
ws.cell(1,2,value="test")
wb.save("test.xlsx")