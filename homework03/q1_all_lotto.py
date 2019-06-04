import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook



wb = Workbook()

# 현재 Active Sheet 얻기
ws = wb.active

ws.cell(1, 1, value='회차')
ws.cell(1, 2, value='번호1')
ws.cell(1, 3, value='번호2')
ws.cell(1, 4, value='번호3')
ws.cell(1, 5, value='번호4')
ws.cell(1, 6, value='번호5')
ws.cell(1, 7, value='번호6')
ws.cell(1, 8, value='보너스 번호')


last_lotto_url = "https://search.daum.net/search?nil_suggest=btn&w=tot&DA=SBC&q=%EB%A1%9C%EB%98%90"
last_lotto_req = requests.get(last_lotto_url)

if last_lotto_req.status_code != requests.codes.ok:
    exit()

last_lotto_html = BeautifulSoup(last_lotto_req.text, 'html.parser')
last_lotto_area = last_lotto_html.select_one('.tit_lotto > .f_red')
last_lotto_num = last_lotto_area.text[:-1]
last_lotto_num = int(last_lotto_num)

print("최종 로또회차:",last_lotto_num)

for lotto_count in range(1, last_lotto_num+1):
    print(lotto_count,"회차 수집")
    lotto_url = "https://search.daum.net/search?w=tot&rtmaxcoll=LOT&DA=LOT&q="+str(lotto_count)+"%ED%9A%8C%EC%B0%A8%20%EB%A1%9C%EB%98%90"
    req = requests.get(lotto_url)

    if req.status_code != requests.codes.ok:
        exit()

    html = BeautifulSoup(req.text, 'html.parser')
    lotto_numbers = html.select('span.img_lotto')


    row_count = lotto_count + 1;
    ws.cell(row_count, 1, value=str(lotto_count) + "회차")
    for index, number in enumerate(lotto_numbers, start=2):
        ws.cell(row_count, index, value=int(number.text))


wb.save("lotto_list.xlsx")
wb.close()